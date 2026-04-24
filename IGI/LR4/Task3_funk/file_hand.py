from __future__ import annotations

import json
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook

from Task3_funk.item import ReportRes


class ResultFileRepository:
    """Work with the append-only text file that stores all reports."""

    def __init__(self, file_path: str | Path = "result_task3.txt"):
        self.file_path = Path(file_path)

    def append_report(self, report: ReportRes) -> None:
        """Append one report to the text file as JSON line."""
        self.file_path.parent.mkdir(parents=True, exist_ok=True)
        with self.file_path.open("a", encoding="utf-8") as f:
            f.write(json.dumps(report.to_dict(), ensure_ascii=False) + "\n")

    def read_all_reports(self) -> list[ReportRes]:
        """Read all reports from the text file."""
        if not self.file_path.exists():
            return []

        reports: list[ReportRes] = []
        with self.file_path.open("r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                reports.append(ReportRes.from_dict(json.loads(line)))
        return reports


class ExcelReportRepository:
    """Append results to an Excel workbook."""

    def __init__(self, file_path: str | Path = "result_task3.xlsx"):
        self.file_path = Path(file_path)

    def _get_workbook(self):
        """Load existing workbook or create a new one."""
        if self.file_path.exists():
            return load_workbook(self.file_path)
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Summary"
        wb.create_sheet("Series")
        return wb

    def _ensure_headers(self, wb):
        """Create headers if sheets are empty."""
        summary = wb["Summary"]
        series = wb["Series"]

        if summary.max_row == 1 and summary["A1"].value is None:
            summary.append([
                "ID", "x", "eps", "iterations",
                "exact_value", "approx_value",
                "mean", "median", "mode", "variance", "std_dev"
            ])

        if series.max_row == 1 and series["A1"].value is None:
            series.append([
                "ID", "x", "n", "term", "partial_sum", "abs_error"
            ])

    def append_report(self, report: ReportRes) -> None:
        """Append report summary and series rows to Excel."""
        wb = self._get_workbook()
        self._ensure_headers(wb)

        summary = wb["Summary"]
        series = wb["Series"]

        report_id = summary.max_row  # next id after header

        stats = report.stats
        summary.append([
            report_id,
            report.x,
            report.eps,
            report.iterations,
            report.exact_value,
            report.approx_value,
            stats.mean_value if stats else None,
            stats.median_value if stats else None,
            stats.mode_value if stats else None,
            stats.variance if stats else None,
            stats.std_dev if stats else None,
        ])

        for row in report.rows:
            series.append([
                report_id,
                report.x,
                row.n,
                row.term,
                row.partial_sum,
                row.abs_error,
            ])

        wb.save(self.file_path)

    def export_all(self, reports: Iterable[ReportRes]) -> None:
        """Rewrite workbook from scratch using all reports."""
        wb = Workbook()
        summary = wb.active
        summary.title = "Summary"
        series = wb.create_sheet("Series")

        summary.append([
            "ID", "x", "eps", "iterations",
            "exact_value", "approx_value",
            "mean", "median", "mode", "variance", "std_dev"
        ])
        series.append([
            "ID", "x", "n", "term", "partial_sum", "abs_error"
        ])

        for report_id, report in enumerate(reports, start=1):
            stats = report.stats
            summary.append([
                report_id,
                report.x,
                report.eps,
                report.iterations,
                report.exact_value,
                report.approx_value,
                stats.mean_value if stats else None,
                stats.median_value if stats else None,
                stats.mode_value if stats else None,
                stats.variance if stats else None,
                stats.std_dev if stats else None,
            ])

            for row in report.rows:
                series.append([
                    report_id,
                    report.x,
                    row.n,
                    row.term,
                    row.partial_sum,
                    row.abs_error,
                ])

        wb.save(self.file_path)